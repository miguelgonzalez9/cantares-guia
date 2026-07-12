# -*- coding: utf-8 -*-
# Cantares — inventario de árboles georreferenciado → data/trees.geojson
#
# Lee el censo "3_Listado Especies Cantares.xlsx" (Duque & Galeano, 2021):
#   - hoja "Especies + Taxonomía": árboles con etiqueta, coordenadas (DDM),
#     altitud, familia, especie, nombre común.
#   - hoja "Especies+Ecología y Morfología": ficha técnica por especie
#     (hábitat, zona de vida, origen, rango altitudinal, morfología, usos,
#     estado UICN).
# Une ambas y escribe una FeatureCollection con descripciones compuestas.
# Los árboles son una CAPA DE REFERENCIA (no editable por el CMS): se cargan
# siempre del archivo estático. Uso:
#   python data_prep/11_trees_inventory_to_geojson.py
import openpyxl, json, re, os

XLSX = os.environ.get('CANTARES_XLSX',
    r'C:\Users\migol\Dropbox\Cantares\info\censos_inventarios\3_Listado Especies Cantares.xlsx')
OUT = os.path.join(os.path.dirname(__file__), '..', 'app', 'public', 'data', 'trees.geojson')

def ddm_to_dd(pos):
    if not pos: return None
    m = re.search(r'([NS])\s*(\d+)\s+(\d+\.?\d*)\s+([EW])\s*(\d+)\s+(\d+\.?\d*)', str(pos))
    if not m: return None
    ns, latd, latm, ew, lngd, lngm = m.groups()
    lat = float(latd) + float(latm)/60.0
    lng = float(lngd) + float(lngm)/60.0
    if ns == 'S': lat = -lat
    if ew == 'W': lng = -lng
    return round(lat, 6), round(lng, 6)

norm = lambda s: re.sub(r'\s+', ' ', str(s).strip().lower()) if s else ''
def clean(s):
    if not s: return ''
    return re.sub(r'\s+', ' ', str(s).strip()).strip(' .;,-')
def cap(s):
    s = clean(s); return s[0].upper() + s[1:] if s else s

FIX_SCI = {
    'Retrophyllum rospigliossi': 'Retrophyllum rospigliosii',
    'Eucaliptus grandis': 'Eucalyptus grandis',
    'Quercus humbolldi': 'Quercus humboldtii',
    'Hedyosmun bonplandianum': 'Hedyosmum bonplandianum',
    'Chrysochlamys colombiano': 'Chrysochlamys colombiana',
}
def sci_display(sp): sp = clean(sp); return FIX_SCI.get(sp, sp)
def is_binomial(sp):
    sp = sci_display(sp)
    if re.search(r'\bsp\.?\d*\b', sp.lower()): return False
    parts = sp.split()
    return len(parts) >= 2 and parts[0][:1].isupper()

wb = openpyxl.load_workbook(XLSX, data_only=True)

# hoja 1: árboles georreferenciados
ws1 = wb['Especies + Taxonomía']
rows1 = list(ws1.iter_rows(values_only=True))
hdr1 = [str(c).strip() if c else '' for c in rows1[0]]
trees = []
for r in rows1[1:]:
    d = dict(zip(hdr1, r))
    dd = ddm_to_dd(d.get('Position'))
    if not dd: continue
    trees.append({'tag': clean(d.get('Name')), 'lat': dd[0], 'lng': dd[1],
        'altitude': clean(d.get('Altitude')), 'familia': clean(d.get('Familia')),
        'especie': clean(d.get('Especie')), 'comun': clean(d.get('Nombre común'))})

# hoja 2: ficha técnica por especie
ws2 = wb['Especies+Ecología y Morfología']
rows2 = list(ws2.iter_rows(values_only=True))
hdr2 = [str(c).strip() if c else '' for c in rows2[0]]
detail = {}
for r in rows2[1:]:
    d = dict(zip(hdr2, r))
    sp = d.get('Especie')
    if not sp: continue
    k = norm(sp)
    if k in detail: continue
    g = lambda *names: next((clean(d.get(n)) for n in names if d.get(n)), '')
    detail[k] = {'familia': g('Familia'), 'habitat': g('Hábitat'),
        'zona': g('Zona de vida o ecosistema'), 'origen': g('Origen ', 'Origen'),
        'rango': g('Rango altitudinal'), 'morfologia': g('Morfologia ', 'Morfologia'),
        'usos': g('Usos'), 'estado': g('Estado De Conservación (Categoría UICN)')}

def compose(t, det):
    parts = []
    fam = (det.get('familia') if det else '') or t.get('familia')
    if det:
        if det.get('morfologia'):
            m = cap(det['morfologia']); parts.append(m + ('' if m.endswith('.') else '.'))
        hz = [x for x in [det.get('habitat', '').lower(), det.get('zona', '')] if x]
        if hz: parts.append('Hábitat: ' + '; '.join(hz).rstrip('.') + '.')
        line = [x for x in [det.get('origen', ''),
            ('entre ' + det['rango'].lstrip('entre ').rstrip('.') + ' de altitud') if det.get('rango') else ''] if x]
        if line: parts.append(cap('. '.join(line)) + '.')
        if det.get('usos'): parts.append('Usos: ' + det['usos'].rstrip('.').lower() + '.')
        if det.get('estado') and det['estado'] not in ('-', 'NA', 'N/A'):
            parts.append('Estado de conservación (UICN): ' + det['estado'] + '.')
    if not parts:
        parts.append(('Árbol nativo del bosque de Cantares, familia %s.' % fam) if fam
                     else 'Árbol del bosque de Cantares.')
    return ' '.join(parts).strip()

feats, seen = [], set()
tree_species = {}   # nombre científico normalizado → datos de especie (para species.json)
for t in trees:
    tag = t['tag']
    if not tag or tag in seen: continue
    seen.add(tag)
    det = detail.get(norm(t['especie']))
    sci = sci_display(t['especie'])
    comun = t['comun']
    title = (comun or sci or 'Árbol').split(' - ')[0].split('/')[0].strip()
    binom = is_binomial(t['especie'])
    # Link con la especie POR NOMBRE CIENTÍFICO (sólo binomios reales; las
    # morfoespecies "sp." no se linkean). La app resuelve el chip por sci name.
    sci_key = norm(sci) if binom else ''
    species_ids = [sci_key] if sci_key else []
    if binom:
        fam = (det.get('familia') if det else '') or t.get('familia')
        tree_species.setdefault(sci_key, {'scientific_name': sci, 'common_name': title,
            'family': fam or None, 'description': compose(t, det)})
    feats.append({'type': 'Feature', 'geometry': {'type': 'Point', 'coordinates': [t['lng'], t['lat']]},
        'properties': {'id': 'arbol_%s' % tag, 'tipo': 'arbol', 'title': title, 'title_en': None,
            'sci': sci if (binom or (sci and 'sp' not in sci.lower())) else None,
            'family': (det.get('familia') if det else '') or t.get('familia') or None,
            'comun_full': comun or None, 'tag': tag, 'altitude': t.get('altitude') or None,
            'description': compose(t, det), 'description_en': None,
            'routes': [], 'species_ids': species_ids, 'photo': None}})

fc = {'type': 'FeatureCollection',
    '_meta': {'note': 'Inventario georreferenciado de árboles de Cantares (censo 2021, Duque & Galeano). Generado por data_prep/11_trees_inventory_to_geojson.py. Puntos tipo "arbol", editables (se fusionan con la nube por id).', 'n': len(feats)},
    'features': feats}
json.dump(fc, open(os.path.normpath(OUT), 'w', encoding='utf-8'), ensure_ascii=False, indent=1)

# --- Añadir a species.json las especies-árbol que aún no estén (por sci name) ---
SP = os.path.join(os.path.dirname(__file__), '..', 'app', 'public', 'data', 'species.json')
sp_doc = json.load(open(os.path.normpath(SP), encoding='utf-8'))
existing = {norm(s.get('scientific_name')) for s in sp_doc['species'] if s.get('scientific_name')}
def slug(sci):
    return re.sub(r'[^a-z0-9]+', '-', norm(sci)).strip('-')
added = 0
for k, s in sorted(tree_species.items()):
    if k in existing: continue
    sp_doc['species'].append({'id': slug(s['scientific_name']), 'group': 'flora', 'status': 'documented',
        'scientific_name': s['scientific_name'], 'common_name': s['common_name'], 'family': s['family'],
        'flagship': False, 'zones': [], 'photo': None,
        'notes': s['description'][:400] if s.get('description') else None, 'source': 'censo_arboles_2021'})
    existing.add(k); added += 1
json.dump(sp_doc, open(os.path.normpath(SP), 'w', encoding='utf-8'), ensure_ascii=False, indent=1)

print('árboles:', len(feats), '| con nombre científico:', sum(1 for f in feats if f['properties']['sci']),
      '| linkeados a especie:', sum(1 for f in feats if f['properties']['species_ids']))
print('especies-árbol únicas (binomios):', len(tree_species), '| añadidas a species.json:', added,
      '| total especies:', len(sp_doc['species']))
